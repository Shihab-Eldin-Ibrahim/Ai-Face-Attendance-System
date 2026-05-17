import cv2
import os
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ==========================================
# PATHS
# ==========================================

dataset_path = "dataset"
trainer_path = "trainer.yml"
excel_file = "attendance.xlsx"

if not os.path.exists(dataset_path):
    os.makedirs(dataset_path)

# ==========================================
# CREATE EXCEL FILE
# ==========================================

if not os.path.exists(excel_file):

    workbook = Workbook()
    sheet = workbook.active

    sheet.title = "Attendance"

    sheet.append(["Name", "Date", "Time"])

    workbook.save(excel_file)

# ==========================================
# FACE DETECTOR
# ==========================================

face_detector = cv2.CascadeClassifier(
    cv2.data.haarcascades +
    "haarcascade_frontalface_default.xml"
)

# ==========================================
# REGISTER FACE
# ==========================================

def register_face():

    name = input("Enter person name: ")

    person_path = os.path.join(dataset_path, name)

    if not os.path.exists(person_path):
        os.makedirs(person_path)

    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)

    count = 0
    max_samples = 100

    print("Collecting face samples...")

    while True:

        ret, frame = cap.read()

        if not ret:
            break

        gray = cv2.cvtColor(
            frame,
            cv2.COLOR_BGR2GRAY
        )

        faces = face_detector.detectMultiScale(
            gray,
            scaleFactor=1.3,
            minNeighbors=5
        )

        for (x, y, w, h) in faces:

            face = gray[y:y+h, x:x+w]

            count += 1

            file_path = os.path.join(
                person_path,
                f"{count}.jpg"
            )

            cv2.imwrite(file_path, face)

            cv2.rectangle(
                frame,
                (x, y),
                (x+w, y+h),
                (0, 255, 0),
                2
            )

            cv2.putText(
                frame,
                f"Samples: {count}/{max_samples}",
                (x, y-10),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.8,
                (0, 255, 0),
                2
            )

        cv2.imshow(
            "Register Face",
            frame
        )

        key = cv2.waitKey(1)

        if key == ord('q') or count >= max_samples:
            break

    cap.release()
    cv2.destroyAllWindows()

    print(f"{name} registered successfully")

# ==========================================
# TRAIN MODEL
# ==========================================

def train_model():

    recognizer = cv2.face.LBPHFaceRecognizer_create()

    faces = []
    labels = []

    label_map = {}

    current_id = 0

    for person_name in os.listdir(dataset_path):

        person_path = os.path.join(
            dataset_path,
            person_name
        )

        if not os.path.isdir(person_path):
            continue

        label_map[current_id] = person_name

        for image_name in os.listdir(person_path):

            image_path = os.path.join(
                person_path,
                image_name
            )

            img = cv2.imread(
                image_path,
                cv2.IMREAD_GRAYSCALE
            )

            if img is None:
                continue

            faces.append(img)
            labels.append(current_id)

        current_id += 1

    recognizer.train(
        faces,
        np.array(labels)
    )

    recognizer.save(trainer_path)

    np.save(
        "labels.npy",
        label_map
    )

    print("Training completed")

# ==========================================
# SAVE ATTENDANCE
# ==========================================

def save_attendance(name):

    workbook = load_workbook(excel_file)
    sheet = workbook.active

    now = datetime.now()

    current_date = now.strftime("%Y-%m-%d")
    current_time = now.strftime("%H:%M:%S")

    already_exists = False

    for row in sheet.iter_rows(min_row=2, values_only=True):

        saved_name, saved_date, _ = row

        if saved_name == name and saved_date == current_date:
            already_exists = True
            break

    if not already_exists:

        sheet.append([
            name,
            current_date,
            current_time
        ])

        workbook.save(excel_file)

        print(f"Attendance saved for {name}")

# ==========================================
# RECOGNITION SYSTEM
# ==========================================

def recognize_faces():

    if not os.path.exists(trainer_path):
        print("Train model first")
        return

    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read(trainer_path)

    label_map = np.load(
        "labels.npy",
        allow_pickle=True
    ).item()

    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)

    while True:

        ret, frame = cap.read()

        if not ret:
            break

        gray = cv2.cvtColor(
            frame,
            cv2.COLOR_BGR2GRAY
        )

        faces = face_detector.detectMultiScale(
            gray,
            scaleFactor=1.3,
            minNeighbors=5
        )

        for (x, y, w, h) in faces:

            face = gray[y:y+h, x:x+w]

            label, confidence = recognizer.predict(face)

            if confidence < 70:

                name = label_map[label]
                color = (0, 255, 0)

                save_attendance(name)

            else:

                name = "Unknown"
                color = (0, 0, 255)

            cv2.rectangle(
                frame,
                (x, y),
                (x+w, y+h),
                color,
                2
            )

            cv2.putText(
                frame,
                name,
                (x, y-10),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.9,
                color,
                2
            )

        cv2.imshow(
            "AI Face Recognition Attendance System",
            frame
        )

        key = cv2.waitKey(1)

        if key == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

# ==========================================
# MENU
# ==========================================

while True:

    print("\n===== AI FACE ATTENDANCE SYSTEM =====")
    print("1. Register New Face")
    print("2. Train Model")
    print("3. Start Recognition")
    print("4. Exit")

    choice = input("Choose option: ")

    if choice == "1":
        register_face()

    elif choice == "2":
        train_model()

    elif choice == "3":
        recognize_faces()

    elif choice == "4":
        break

    else:
        print("Invalid option")